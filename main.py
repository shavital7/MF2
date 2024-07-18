import os
from openpyxl import load_workbook
from openpyxl import Workbook


def add_cal_transactions(file_path, merged_sheet):
    wb = load_workbook(file_path)
    sheet = wb[wb.sheetnames[0]]

    for row in sheet.iter_rows(values_only=True, min_row=3):
        merged_sheet.append(row)

def merge_excel_files(input_folder, output_file):
    merged_workbook = Workbook()
    merged_sheet = merged_workbook.active

    for dir in os.listdir(input_folder):
        if (dir == "Cal"):
            cal_path = os.path.join(input_folder, "Cal")
            for cal_file in os.listdir(cal_path):
                if cal_file.endswith(".xlsx"):
                    add_cal_transactions(os.path.join(cal_path, cal_file), merged_sheet)

    # Save the merged workbook to the specified output file
    merged_workbook.save(output_file)
    print(f"Merged Excel file saved as {output_file}")

# Example usage:
if __name__ == "__main__":
    input_folder = "Reports"  # Replace with the path to your input folder containing Excel files
    output_file = "Reports/output.xlsx"  # Replace with the desired name of the output Excel file

    print("starting")
    merge_excel_files(input_folder, output_file)